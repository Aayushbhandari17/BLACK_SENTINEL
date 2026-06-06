from openpyxl import load_workbook

def extract_text(file_path):
    workbook = load_workbook(file_path, data_only=True)

    rows = []

    for sheet in workbook.worksheets:
        rows.append(f"[SHEET:{sheet.title}]")

        headers = []
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if r_idx == 0:
                headers = [str(cell).strip() if cell is not None else f"col{i}" for i, cell in enumerate(row)]
                values = [h for h in headers if h]
                if values:
                    rows.append(" | ".join(values))
                continue

            row_strs = []
            for c_idx, cell in enumerate(row):
                if cell is not None:
                    header = headers[c_idx] if c_idx < len(headers) else f"col{c_idx}"
                    row_strs.append(f"{header}: {cell}")

            if row_strs:
                rows.append(" | ".join(row_strs))

    return "\n".join(rows)

class XLSXParser:
    def parse(self, file_path):
        return extract_text(file_path)